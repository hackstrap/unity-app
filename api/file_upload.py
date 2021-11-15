import datetime
import json
import math
from datetime import date, datetime

import arrow as arw
import numpy as np
import pandas as pd
import pytz
import requests
from dateutil import parser
from dateutil.relativedelta import relativedelta
from flask import Blueprint, jsonify, request
from flask_cors import CORS, cross_origin
from openpyxl import Workbook, load_workbook
from requests.api import get
from requests.exceptions import HTTPError
from rich.console import Console
from werkzeug.utils import secure_filename

console = Console()

ALLOWED_EXCEL_EXTENSIONS = {"xlsx"}

from utils.default_data import default_portfolio

base_url = "https://blink.hackstrap.com/"

local_url = "http://127.0.0.1:5000/"

PREFIX = "Bearer"


def get_token(header):
    bearer, _, token = header.partition(" ")
    if bearer != PREFIX:
        raise ValueError("Invalid token")

    return token


api_file_upload = Blueprint("api_file_upload", __name__)
CORS(api_file_upload)


@api_file_upload.route("/unity/v1/api/file_upload", methods=["POST"])
def file_upload():
    page = 0
    page_size = 12
    startup_id = request.args.get("startup_id")
    year = int
    month = int
    header = request.headers.get("Authorization")
    access_token = get_token(header)

    if request.method == "POST":
        # check if the post request has the file part
        if "file" not in request.files:
            return "No file part"
        else:
            file = request.files["file"]
            if file.filename == "":
                return "No selected file"
            else:
                file = request.files["file"]
                filename = secure_filename(file.filename)
                file_extension = (
                    filename.rsplit(".", 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS
                )
                if file_extension == False:
                    return "Only xlsx file extension allowed"
                else:
                    file = request.files["file"]
                    # print(request.files)
                    print(startup_id)
                    filename = secure_filename(file.filename)
                    print(filename)

                    wb = load_workbook(request.files["file"])
                    print(wb.sheetnames)
                    sheet_list = [
                        "Info",
                        "Revenue",
                        "Users",
                        "Expense",
                        "Opex Expenses",
                        "Employee",
                        "Product",
                    ]

                    # check if sheetnames are in the sheet_list
                    for sheet in wb.sheetnames:
                        if sheet not in sheet_list:
                            return "Please upload the correct sheets"
                        else:
                            for index in range(len(wb.sheetnames)):
                                print(index, sheet_list[index])
                                # loop through the sheets
                                if sheet_list[index] == "Info":

                                    info_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Info",
                                        dtype={"Format": str, "startup_id": str},
                                    )
                                    # print(info_data)
                                    # print(
                                    #     "Sheet format: " + str(info_data["Format"][0])
                                    # )
                                    # print(
                                    #     "Startup Id: " + str(info_data["startup_id"][0])
                                    # )
                                    print("Info Sheet Processed")
                                elif sheet_list[index] == "Revenue":
                                    # Check for revenue data
                                    revenue_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Revenue",
                                        usecols=["Table Name", "Year"],
                                        dtype={
                                            "Table Name": str,
                                        },
                                    )
                                    revenue_data["Year"] = revenue_data["Year"].fillna(
                                        0
                                    )
                                    revenue_data = revenue_data.astype({"Year": "int"})
                                    year = revenue_data["Year"][0]

                                    # print(revenue_data)
                                    # print(revenue_data["Table Name"][0])
                                    # print(revenue_data["Year"][0])

                                    revenue_data_table = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Revenue",
                                        usecols=[
                                            "Table Column Name",
                                            "January",
                                            "February",
                                            "March",
                                            "April",
                                            "May",
                                            "June",
                                            "July",
                                            "August",
                                            "September",
                                            "October",
                                            "November",
                                            "December",
                                        ],
                                        dtype={
                                            "Table Column Name": str,
                                        },
                                    )

                                    revenue_data_table = revenue_data_table.where(
                                        revenue_data_table.notnull(), None
                                    )

                                    # print(revenue_data_table)

                                    revenue_data_table = revenue_data_table.set_index(
                                        "Table Column Name"
                                    )
                                    revenue_data_table.index.names = [None]
                                    revenue_data_table = revenue_data_table.transpose()

                                    revenue_data_table = pd.DataFrame(
                                        revenue_data_table, dtype=object
                                    )
                                    revenue_data_table.index = [
                                        0,
                                        1,
                                        2,
                                        3,
                                        4,
                                        5,
                                        6,
                                        7,
                                        8,
                                        9,
                                        10,
                                        11,
                                    ]

                                    revenue_data_table = revenue_data_table.where(
                                        revenue_data_table.notnull(), None
                                    )

                                    # print(revenue_data_table)

                                    revenue_data_table[
                                        "Total MRR"
                                    ] = revenue_data_table["Total MRR"].astype(float)
                                    revenue_data_table[
                                        "Total New MRR"
                                    ] = revenue_data_table["Total New MRR"].astype(
                                        float
                                    )
                                    revenue_data_table[
                                        "Total Non-Recurring Revenue"
                                    ] = revenue_data_table[
                                        "Total Non-Recurring Revenue"
                                    ].astype(
                                        float
                                    )

                                    revenue_data_table = revenue_data_table.where(
                                        revenue_data_table.notnull(), None
                                    )

                                    # Check if revenue data is available for that year
                                    revenue_table_result = requests.get(
                                        base_url
                                        + "v1/revenue?"
                                        + "page={}&page_size={}&startup_id={}&year={}".format(
                                            page, page_size, startup_id, year
                                        ),
                                        headers={
                                            "Content-Type": "application/json",
                                            "Authorization": "Bearer {}".format(
                                                access_token
                                            ),
                                        },
                                    )
                                    # If revenue data is not available for that year
                                    if len(revenue_table_result.text) < 4:

                                        json_data = {
                                            "total_revenue": None,
                                            "total_mrr": 0,
                                            "total_new_mrr": 0,
                                            "total_non_recurring_revenue": 0,
                                            "startup_id": "",
                                            "month": 0,
                                            "year": 0,
                                        }
                                        total_mrr_list = revenue_data_table[
                                            "Total MRR"
                                        ].tolist()

                                        total_new_mrr_list = revenue_data_table[
                                            "Total New MRR"
                                        ].tolist()

                                        total_non_recurring_revenue_list = (
                                            revenue_data_table[
                                                "Total Non-Recurring Revenue"
                                            ].tolist()
                                        )

                                        json_revenue_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        for i in json_revenue_range:

                                            json_data["startup_id"] = str(startup_id)
                                            json_data["month"] = int(i + 1)
                                            json_data["year"] = int(year)
                                            json_data["total_mrr"] = total_mrr_list[i]
                                            json_data[
                                                "total_new_mrr"
                                            ] = total_new_mrr_list[i]
                                            json_data[
                                                "total_non_recurring_revenue"
                                            ] = total_non_recurring_revenue_list[i]

                                            # print(json_data)

                                            r = requests.post(
                                                base_url + "v1/revenue/",
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )

                                            # print(base_url + "v1/revenue/{}".format(id))
                                            # print(json_data)
                                            # print(r.text)

                                    # If revenue data is available for that year
                                    else:

                                        revenue_table_result = json.loads(
                                            revenue_table_result.text
                                        )
                                        revenue_table_result = pd.DataFrame(
                                            revenue_table_result
                                        )

                                        try:
                                            revenue_table_result = (
                                                revenue_table_result.sort_values(
                                                    by="month", ascending=True
                                                )
                                            )
                                        except:
                                            # print("Block1")
                                            return jsonify([])

                                        revenue_table_result.index = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        revenue_table_result = (
                                            revenue_table_result.where(
                                                revenue_table_result.notnull(), None
                                            )
                                        )

                                        # print(revenue_table_result)
                                        # print(
                                        #     type(revenue_table_result["total_mrr"][0])
                                        # )
                                        # print(
                                        #     type(revenue_table_result["total_mrr"][1])
                                        # )

                                        # print(revenue_table_result["total_mrr"])
                                        # print(revenue_data_table["Total MRR"])

                                        revenue_table_result[
                                            "total_mrr"
                                        ] = revenue_data_table["Total MRR"]
                                        revenue_table_result[
                                            "total_new_mrr"
                                        ] = revenue_data_table["Total New MRR"]

                                        revenue_table_result[
                                            "total_non_recurring_revenue"
                                        ] = revenue_data_table[
                                            "Total Non-Recurring Revenue"
                                        ]

                                        revenue_data = revenue_table_result

                                        # print(revenue_data)
                                        # print(type(revenue_data["total_mrr"][0]))
                                        # print(type(revenue_data["total_mrr"][1]))

                                        revenue_data = revenue_data.to_dict(
                                            orient="records"
                                        )

                                        revenue_data = json.dumps(revenue_data)

                                        revenue_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        id_list = revenue_table_result["_id"].tolist()

                                        total_mrr_list = revenue_table_result[
                                            "total_mrr"
                                        ].tolist()

                                        total_new_mrr_list = revenue_table_result[
                                            "total_new_mrr"
                                        ].tolist()

                                        total_non_recurring_revenue_list = (
                                            revenue_table_result[
                                                "total_non_recurring_revenue"
                                            ].tolist()
                                        )

                                        json_data = {
                                            "total_mrr": 0,
                                            "total_new_mrr": 0,
                                            "total_non_recurring_revenue": 0,
                                        }

                                        for i in revenue_range:
                                            # print(revenue_table_result["_id"][i])
                                            # print(type(revenue_table_result["total_revenue"][i]))

                                            json_data["total_mrr"] = total_mrr_list[i]
                                            json_data[
                                                "total_new_mrr"
                                            ] = total_new_mrr_list[i]
                                            json_data[
                                                "total_non_recurring_revenue"
                                            ] = total_non_recurring_revenue_list[i]

                                            id = id_list[i]
                                            # print(id)
                                            # print(json_data)
                                            # print("v1/revenue/{}".format(id))

                                            r = requests.put(
                                                base_url + "v1/revenue/{}".format(id),
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )

                                            # print(base_url + "v1/revenue/{}".format(id))
                                            # print(json_data)
                                            # print(r.text)

                                    print("Revenue Sheet Processed")

                                elif sheet_list[index] == "Users":
                                    # check for users data
                                    users_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Users",
                                        usecols=["Table Name", "Year"],
                                        dtype={
                                            "Table Name": str,
                                        },
                                    )
                                    users_data["Year"] = users_data["Year"].fillna(0)
                                    users_data = users_data.astype({"Year": "int"})

                                    year = users_data["Year"][0]
                                    # print(users_data)
                                    # print(users_data["Table Name"][0])
                                    # print(users_data["Year"][0])

                                    users_data_table = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Users",
                                        usecols=[
                                            "Table Column Name",
                                            "January",
                                            "February",
                                            "March",
                                            "April",
                                            "May",
                                            "June",
                                            "July",
                                            "August",
                                            "September",
                                            "October",
                                            "November",
                                            "December",
                                        ],
                                        dtype={
                                            "Table Column Name": str,
                                        },
                                    )
                                    users_data_table = users_data_table.where(
                                        users_data_table.notnull(), None
                                    )

                                    # print(
                                    #     type(users_data_table["Table Column Name"][0])
                                    # )
                                    # print(type(users_data_table["January"][0]))
                                    # print(type(users_data_table["December"][0]))
                                    # print(users_data_table)
                                    users_data_table = users_data_table.set_index(
                                        "Table Column Name"
                                    )
                                    users_data_table.index.names = [None]
                                    users_data_table = users_data_table.transpose()

                                    users_data_table = pd.DataFrame(
                                        users_data_table, dtype=object
                                    )
                                    users_data_table.index = [
                                        0,
                                        1,
                                        2,
                                        3,
                                        4,
                                        5,
                                        6,
                                        7,
                                        8,
                                        9,
                                        10,
                                        11,
                                    ]
                                    # print(users_data_table)

                                    users_data_table[
                                        "Total Registered Users"
                                    ] = users_data_table[
                                        "Total Registered Users"
                                    ].astype(
                                        float
                                    )
                                    users_data_table[
                                        "Total Monthly Active Users"
                                    ] = users_data_table[
                                        "Total Monthly Active Users"
                                    ].astype(
                                        float
                                    )
                                    users_data_table[
                                        "Total Customers at the beginning of the month"
                                    ] = users_data_table[
                                        "Total Customers at the beginning of the month"
                                    ].astype(
                                        float
                                    )
                                    users_data_table[
                                        "Total New Customers Acquired"
                                    ] = users_data_table[
                                        "Total New Customers Acquired"
                                    ].astype(
                                        float
                                    )
                                    users_data_table[
                                        "Total Customers Churned"
                                    ] = users_data_table[
                                        "Total Customers Churned"
                                    ].astype(
                                        float
                                    )

                                    users_data_table = users_data_table.where(
                                        users_data_table.notnull(), None
                                    )

                                    # check if users data is available for that year
                                    users_table_result = requests.get(
                                        base_url
                                        + "v1/users?"
                                        + "page={}&page_size={}&startup_id={}&year={}".format(
                                            page, page_size, startup_id, year
                                        ),
                                        headers={
                                            "Content-Type": "application/json",
                                            "Authorization": "Bearer {}".format(
                                                access_token
                                            ),
                                        },
                                    )

                                    # If users data is not available for that year
                                    if len(users_table_result.text) < 4:

                                        json_data = {
                                            "startup_id": None,
                                            "month": None,
                                            "year": None,
                                            "total_customers": None,
                                            "total_customers_at_beginning_of_month": None,
                                            "total_customers_churned": None,
                                            "total_monthly_active_users": None,
                                            "total_new_customers_acquired": None,
                                            "total_registered_users": None,
                                        }

                                        total_customers_at_beginning_of_month_list = users_data_table[
                                            "Total Customers at the beginning of the month"
                                        ].tolist()

                                        total_customers_churned_list = users_data_table[
                                            "Total Customers Churned"
                                        ].tolist()
                                        total_monthly_active_users_list = (
                                            users_data_table[
                                                "Total Monthly Active Users"
                                            ].tolist()
                                        )

                                        total_new_customers_acquired_list = (
                                            users_data_table[
                                                "Total New Customers Acquired"
                                            ].tolist()
                                        )

                                        total_registered_users_list = users_data_table[
                                            "Total Registered Users"
                                        ].tolist()

                                        json_users_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]
                                        for i in json_users_range:

                                            json_data["startup_id"] = str(startup_id)
                                            json_data["month"] = int(i + 1)
                                            json_data["year"] = int(year)
                                            json_data[
                                                "total_customers_at_beginning_of_month"
                                            ] = total_customers_at_beginning_of_month_list[
                                                i
                                            ]
                                            json_data[
                                                "total_customers_churned"
                                            ] = total_customers_churned_list[i]
                                            json_data[
                                                "total_monthly_active_users"
                                            ] = total_monthly_active_users_list[i]
                                            json_data[
                                                "total_new_customers_acquired"
                                            ] = total_new_customers_acquired_list[i]
                                            json_data[
                                                "total_registered_users"
                                            ] = total_registered_users_list[i]

                                            # print(json_data)

                                            r = requests.post(
                                                base_url + "v1/users/",
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)
                                            # print(r.text)

                                    # If user data is available for that year
                                    else:

                                        users_table_result = json.loads(
                                            users_table_result.text
                                        )

                                        users_table_result = pd.DataFrame(
                                            users_table_result
                                        )
                                        try:
                                            users_table_result = (
                                                users_table_result.sort_values(
                                                    by="month", ascending=True
                                                )
                                            )
                                        except:
                                            # print("Block2")
                                            return jsonify([])

                                        users_table_result.index = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        users_table_result = users_table_result.where(
                                            users_table_result.notnull(), None
                                        )

                                        users_table_result[
                                            "total_customers_at_beginning_of_month"
                                        ] = users_data_table[
                                            "Total Customers at the beginning of the month"
                                        ]
                                        users_table_result[
                                            "total_customers_churned"
                                        ] = users_data_table["Total Customers Churned"]
                                        users_table_result[
                                            "total_monthly_active_users"
                                        ] = users_data_table[
                                            "Total Monthly Active Users"
                                        ]
                                        users_table_result[
                                            "total_new_customers_acquired"
                                        ] = users_data_table[
                                            "Total New Customers Acquired"
                                        ]
                                        users_table_result[
                                            "total_registered_users"
                                        ] = users_data_table["Total Registered Users"]

                                        users_data = users_table_result

                                        # print(users_data)

                                        users_data = users_data.to_dict(
                                            orient="records"
                                        )

                                        users_data = json.dumps(users_data)
                                        users_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        id_list = users_table_result["_id"].tolist()

                                        total_customers_at_beginning_of_month_list = users_data_table[
                                            "Total Customers at the beginning of the month"
                                        ].tolist()

                                        total_customers_churned_list = users_data_table[
                                            "Total Customers Churned"
                                        ].tolist()

                                        total_monthly_active_users_list = (
                                            users_data_table[
                                                "Total Monthly Active Users"
                                            ].tolist()
                                        )

                                        total_new_customers_acquired_list = (
                                            users_data_table[
                                                "Total New Customers Acquired"
                                            ].tolist()
                                        )

                                        total_registered_users_list = users_data_table[
                                            "Total Registered Users"
                                        ].tolist()

                                        json_data = {
                                            "total_customers_at_beginning_of_month": None,
                                            "total_customers_churned": None,
                                            "total_monthly_active_users": None,
                                            "total_new_customers_acquired": None,
                                            "total_registered_users": None,
                                        }

                                        for i in users_range:

                                            json_data[
                                                "total_customers_at_beginning_of_month"
                                            ] = total_customers_at_beginning_of_month_list[
                                                i
                                            ]
                                            json_data[
                                                "total_customers_churned"
                                            ] = total_customers_churned_list[i]
                                            json_data[
                                                "total_monthly_active_users"
                                            ] = total_monthly_active_users_list[i]
                                            json_data[
                                                "total_new_customers_acquired"
                                            ] = total_new_customers_acquired_list[i]
                                            json_data[
                                                "total_registered_users"
                                            ] = total_registered_users_list[i]
                                            id = id_list[i]
                                            # print(id)
                                            # print(json_data)

                                            r = requests.put(
                                                base_url + "v1/users/{}".format(id),
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)

                                    print("Users Sheet Processed")

                                elif sheet_list[index] == "Expense":
                                    # check for expense data
                                    expense_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Expense",
                                        usecols=["Table Name", "Year"],
                                        dtype={
                                            "Table Name": str,
                                        },
                                    )
                                    expense_data["Year"] = expense_data["Year"].fillna(
                                        0
                                    )
                                    expense_data = expense_data.astype({"Year": "int"})
                                    year = expense_data["Year"][0]

                                    # print(expense_data)
                                    # print(expense_data["Table Name"][0])
                                    # print(expense_data["Year"][0])

                                    expense_data_table = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Expense",
                                        usecols=[
                                            "Table Column Name",
                                            "January",
                                            "February",
                                            "March",
                                            "April",
                                            "May",
                                            "June",
                                            "July",
                                            "August",
                                            "September",
                                            "October",
                                            "November",
                                            "December",
                                        ],
                                        dtype={
                                            "Table Column Name": str,
                                        },
                                    )

                                    expense_data_table = expense_data_table.where(
                                        expense_data_table.notnull(), None
                                    )
                                    # print(expense_data_table)

                                    expense_data_table = expense_data_table.set_index(
                                        "Table Column Name"
                                    )
                                    expense_data_table.index.names = [None]
                                    expense_data_table = expense_data_table.transpose()
                                    expense_data_table = pd.DataFrame(
                                        expense_data_table, dtype=object
                                    )
                                    expense_data_table.index = [
                                        0,
                                        1,
                                        2,
                                        3,
                                        4,
                                        5,
                                        6,
                                        7,
                                        8,
                                        9,
                                        10,
                                        11,
                                    ]
                                    # print(expense_data_table)

                                    expense_data_table[
                                        "Total Payroll-Support"
                                    ] = expense_data_table[
                                        "Total Payroll-Support"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Software & Tools-Support"
                                    ] = expense_data_table[
                                        "Software & Tools-Support"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Hosting-Service Delivery"
                                    ] = expense_data_table[
                                        "Hosting-Service Delivery"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Direct Material Costs"
                                    ] = expense_data_table[
                                        "Direct Material Costs"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Direct Labor Costs"
                                    ] = expense_data_table["Direct Labor Costs"].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Manufacturing Overhead"
                                    ] = expense_data_table[
                                        "Manufacturing Overhead"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Net WIP Inventory"
                                    ] = expense_data_table["Net WIP Inventory"].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Net Finished Goods Inventory"
                                    ] = expense_data_table[
                                        "Net Finished Goods Inventory"
                                    ].astype(
                                        float
                                    )
                                    expense_data_table[
                                        "Total Other COGS"
                                    ] = expense_data_table["Total Other COGS"].astype(
                                        float
                                    )

                                    expense_data_table = expense_data_table.where(
                                        expense_data_table.notnull(), None
                                    )

                                    # check if expense data is available for that year

                                    expense_table_result = requests.get(
                                        base_url
                                        + "v1/expense?"
                                        + "page={}&page_size={}&startup_id={}&year={}".format(
                                            page, page_size, startup_id, year
                                        ),
                                        headers={
                                            "Content-Type": "application/json",
                                            "Authorization": "Bearer {}".format(
                                                access_token
                                            ),
                                        },
                                    )

                                    # If expense data is not available for that year
                                    if len(expense_table_result.text) < 4:

                                        json_data = {
                                            "direct_labor_costs": 0,
                                            "direct_material_costs": 0,
                                            "hosting_service_delivery": 0,
                                            "manufacturing_overhead": 0,
                                            "month": 0,
                                            "net_finished_goods_inventory": 0,
                                            "net_wip_inventory": 0,
                                            "software_and_tools_support": 0,
                                            "startup_id": "",
                                            "total_cogs": None,
                                            "total_cost_of_goods_manufactured": None,
                                            "total_customer_support_expenses": None,
                                            "total_other_cogs": 0,
                                            "total_payroll_support": 0,
                                            "total_service_delivery_expenses": None,
                                            "year": 0,
                                        }

                                        total_payroll_support_list = expense_data_table[
                                            "Total Payroll-Support"
                                        ].to_list()
                                        software_and_tools_support_list = (
                                            expense_data_table[
                                                "Software & Tools-Support"
                                            ].to_list()
                                        )
                                        hosting_service_delivery_list = (
                                            expense_data_table[
                                                "Hosting-Service Delivery"
                                            ].to_list()
                                        )
                                        direct_material_costs_list = expense_data_table[
                                            "Direct Material Costs"
                                        ].to_list()
                                        direct_labor_costs_list = expense_data_table[
                                            "Direct Labor Costs"
                                        ].to_list()
                                        manufacturing_overhead_list = (
                                            expense_data_table[
                                                "Manufacturing Overhead"
                                            ].to_list()
                                        )
                                        net_wip_inventory_list = expense_data_table[
                                            "Net WIP Inventory"
                                        ].to_list()
                                        net_finished_goods_inventory_list = (
                                            expense_data_table[
                                                "Net Finished Goods Inventory"
                                            ].to_list()
                                        )
                                        total_other_cogs_list = expense_data_table[
                                            "Total Other COGS"
                                        ].to_list()

                                        json_expense_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        for i in json_expense_range:
                                            json_data["startup_id"] = str(startup_id)
                                            json_data["month"] = int(i + 1)
                                            json_data["year"] = int(year)

                                            json_data[
                                                "total_payroll_support_list"
                                            ] = total_payroll_support_list[i]
                                            json_data[
                                                "software_and_tools_support_list"
                                            ] = software_and_tools_support_list[i]
                                            json_data[
                                                "hosting_service_delivery_list"
                                            ] = hosting_service_delivery_list[i]
                                            json_data[
                                                "direct_material_costs_list"
                                            ] = direct_material_costs_list[i]
                                            json_data[
                                                "direct_labor_costs_list"
                                            ] = direct_labor_costs_list[i]
                                            json_data[
                                                "manufacturing_overhead_list"
                                            ] = manufacturing_overhead_list[i]
                                            json_data[
                                                "net_wip_inventory_list"
                                            ] = net_wip_inventory_list[i]
                                            json_data[
                                                "net_finished_goods_inventory_list"
                                            ] = net_finished_goods_inventory_list[i]
                                            json_data[
                                                "total_other_cogs_list"
                                            ] = total_other_cogs_list[i]

                                            # print(json_data)

                                            r = requests.post(
                                                base_url + "v1/expense/",
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)
                                    # If expense data is available for that year
                                    else:
                                        expense_table_result = json.loads(
                                            expense_table_result.text
                                        )
                                        expense_table_result = pd.DataFrame(
                                            expense_table_result
                                        )

                                        try:
                                            expense_table_result = (
                                                expense_table_result.sort_values(
                                                    by=["month"], ascending=True
                                                )
                                            )
                                        except:
                                            # print("Block3")
                                            return jsonify([])

                                        expense_table_result.index = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        expense_table_result = (
                                            expense_table_result.where(
                                                expense_table_result.notnull(), None
                                            )
                                        )

                                        # print(expense_table_result)

                                        expense_table_result[
                                            "total_payroll_support"
                                        ] = expense_data_table["Total Payroll-Support"]
                                        expense_table_result[
                                            "software_and_tools_support"
                                        ] = expense_data_table[
                                            "Software & Tools-Support"
                                        ]
                                        expense_table_result[
                                            "hosting_service_delivery"
                                        ] = expense_data_table[
                                            "Hosting-Service Delivery"
                                        ]
                                        expense_table_result[
                                            "direct_material_costs"
                                        ] = expense_data_table["Direct Material Costs"]
                                        expense_table_result[
                                            "direct_labor_costs"
                                        ] = expense_data_table["Direct Labor Costs"]
                                        expense_table_result[
                                            "manufacturing_overhead"
                                        ] = expense_data_table["Manufacturing Overhead"]
                                        expense_table_result[
                                            "net_wip_inventory"
                                        ] = expense_data_table["Net WIP Inventory"]
                                        expense_table_result[
                                            "net_finished_goods_inventory"
                                        ] = expense_data_table[
                                            "Net Finished Goods Inventory"
                                        ]
                                        expense_table_result[
                                            "total_other_cogs"
                                        ] = expense_data_table["Total Other COGS"]

                                        expense_data = expense_table_result

                                        expense_data = expense_data.to_dict(
                                            orient="records"
                                        )
                                        expense_data = json.dumps(expense_data)

                                        expense_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        id_list = expense_table_result["_id"].tolist()

                                        total_payroll_support_list = expense_data_table[
                                            "Total Payroll-Support"
                                        ].to_list()

                                        software_and_tools_support_list = (
                                            expense_data_table[
                                                "Software & Tools-Support"
                                            ].to_list()
                                        )
                                        hosting_service_delivery_list = (
                                            expense_data_table[
                                                "Hosting-Service Delivery"
                                            ].to_list()
                                        )
                                        direct_material_costs_list = expense_data_table[
                                            "Direct Material Costs"
                                        ].to_list()

                                        direct_labor_costs_list = expense_data_table[
                                            "Direct Labor Costs"
                                        ].to_list()

                                        manufacturing_overhead_list = (
                                            expense_data_table[
                                                "Manufacturing Overhead"
                                            ].to_list()
                                        )
                                        net_wip_inventory_list = expense_data_table[
                                            "Net WIP Inventory"
                                        ].to_list()

                                        net_finished_goods_inventory_list = (
                                            expense_data_table[
                                                "Net Finished Goods Inventory"
                                            ].to_list()
                                        )
                                        total_other_cogs_list = expense_data_table[
                                            "Total Other COGS"
                                        ].to_list()

                                        json_data = {
                                            "direct_labor_costs": 0,
                                            "direct_material_costs": 0,
                                            "hosting_service_delivery": 0,
                                            "manufacturing_overhead": 0,
                                            "net_finished_goods_inventory": 0,
                                            "net_wip_inventory": 0,
                                            "software_and_tools_support": 0,
                                            "total_other_cogs": 0,
                                            "total_payroll_support": 0,
                                        }

                                        for i in expense_range:

                                            json_data[
                                                "total_payroll_support_list"
                                            ] = total_payroll_support_list[i]
                                            json_data[
                                                "software_and_tools_support_list"
                                            ] = software_and_tools_support_list[i]
                                            json_data[
                                                "hosting_service_delivery_list"
                                            ] = hosting_service_delivery_list[i]
                                            json_data[
                                                "direct_material_costs_list"
                                            ] = direct_material_costs_list[i]
                                            json_data[
                                                "direct_labor_costs_list"
                                            ] = direct_labor_costs_list[i]
                                            json_data[
                                                "manufacturing_overhead_list"
                                            ] = manufacturing_overhead_list[i]
                                            json_data[
                                                "net_wip_inventory_list"
                                            ] = net_wip_inventory_list[i]
                                            json_data[
                                                "net_finished_goods_inventory_list"
                                            ] = net_finished_goods_inventory_list[i]
                                            json_data[
                                                "total_other_cogs_list"
                                            ] = total_other_cogs_list[i]

                                            id = id_list[i]
                                            # print(id)
                                            # print(json_data)

                                            r = requests.put(
                                                base_url + "v1/expense/{}".format(id),
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)

                                    print("Expense Sheet Processed")

                                elif sheet_list[index] == "Opex Expenses":
                                    # check for Opex Expenses data
                                    opex_expenses_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Opex Expenses",
                                        usecols=["Table Name", "Year"],
                                        dtype={
                                            "Table Name": str,
                                        },
                                    )
                                    opex_expenses_data["Year"] = opex_expenses_data[
                                        "Year"
                                    ].fillna(0)
                                    opex_expenses_data = opex_expenses_data.astype(
                                        {"Year": "int"}
                                    )
                                    year = opex_expenses_data["Year"][0]

                                    # print(opex_expenses_data)
                                    # print(opex_expenses_data["Table Name"][0])
                                    # print(opex_expenses_data["Year"][0])

                                    opex_expenses_data_table = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Opex Expenses",
                                        usecols=[
                                            "Table Column Name",
                                            "January",
                                            "February",
                                            "March",
                                            "April",
                                            "May",
                                            "June",
                                            "July",
                                            "August",
                                            "September",
                                            "October",
                                            "November",
                                            "December",
                                        ],
                                        dtype={
                                            "Table Column Name": str,
                                        },
                                    )

                                    opex_expenses_data_table = (
                                        opex_expenses_data_table.where(
                                            opex_expenses_data_table.notnull(), None
                                        )
                                    )

                                    # print(opex_expenses_data_table)

                                    opex_expenses_data_table = (
                                        opex_expenses_data_table.set_index(
                                            "Table Column Name"
                                        )
                                    )
                                    opex_expenses_data_table.index.names = [None]
                                    opex_expenses_data_table = (
                                        opex_expenses_data_table.transpose()
                                    )

                                    opex_expenses_data_table = pd.DataFrame(
                                        opex_expenses_data_table, dtype=object
                                    )

                                    opex_expenses_data_table.index = [
                                        0,
                                        1,
                                        2,
                                        3,
                                        4,
                                        5,
                                        6,
                                        7,
                                        8,
                                        9,
                                        10,
                                        11,
                                    ]

                                    opex_expenses_data_table = (
                                        opex_expenses_data_table.where(
                                            opex_expenses_data_table.notnull(), None
                                        )
                                    )
                                    # print(opex_expenses_data_table)

                                    opex_expenses_data_table[
                                        "Total General and Administrative Expenses"
                                    ] = opex_expenses_data_table[
                                        "Total General and Administrative Expenses"
                                    ].astype(
                                        float
                                    )
                                    opex_expenses_data_table[
                                        "Total Sales and Marketing Expenses"
                                    ] = opex_expenses_data_table[
                                        "Total Sales and Marketing Expenses"
                                    ].astype(
                                        float
                                    )
                                    opex_expenses_data_table[
                                        "Total Research and Development Expenses"
                                    ] = opex_expenses_data_table[
                                        "Total Research and Development Expenses"
                                    ].astype(
                                        float
                                    )
                                    opex_expenses_data_table[
                                        "Total Other Expenses"
                                    ] = opex_expenses_data_table[
                                        "Total Other Expenses"
                                    ].astype(
                                        float
                                    )

                                    opex_expenses_data_table = (
                                        opex_expenses_data_table.where(
                                            opex_expenses_data_table.notnull(), None
                                        )
                                    )

                                    # Check if opex expenses data is available for that year
                                    opex_expenses_table_result = requests.get(
                                        base_url
                                        + "v1/opex?"
                                        + "page={}&page_size={}&startup_id={}&year={}".format(
                                            page, page_size, startup_id, year
                                        ),
                                        headers={
                                            "Content-Type": "application/json",
                                            "Authorization": "Bearer {}".format(
                                                access_token
                                            ),
                                        },
                                    )

                                    # If opex expenses data is not available for that year
                                    if len(opex_expenses_table_result.text) < 4:

                                        json_data = {
                                            "startup_id": "",
                                            "month": 0,
                                            "year": 0,
                                            "total_opex_expenses": None,
                                            "total_general_and_administrative_expenses": 0,
                                            "total_sales_and_marketing_expenses": 0,
                                            "total_research_and_development_expenses": 0,
                                            "total_other_expenses": 0,
                                        }

                                        total_general_and_administrative_expenses_list = opex_expenses_data_table[
                                            "Total General and Administrative Expenses"
                                        ].tolist()
                                        total_sales_and_marketing_expenses_list = (
                                            opex_expenses_data_table[
                                                "Total Sales and Marketing Expenses"
                                            ].tolist()
                                        )
                                        total_research_and_development_expenses_list = opex_expenses_data_table[
                                            "Total Research and Development Expenses"
                                        ].tolist()
                                        total_other_expenses_list = (
                                            opex_expenses_data_table[
                                                "Total Other Expenses"
                                            ].tolist()
                                        )

                                        json_opex_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        for i in json_opex_range:

                                            json_data["startup_id"] = str(startup_id)
                                            json_data["month"] = int(i + 1)
                                            json_data["year"] = int(year)

                                            json_data[
                                                "total_general_and_administrative_expenses"
                                            ] = total_general_and_administrative_expenses_list[
                                                i
                                            ]
                                            json_data[
                                                "total_sales_and_marketing_expenses"
                                            ] = total_sales_and_marketing_expenses_list[
                                                i
                                            ]

                                            json_data[
                                                "total_research_and_development_expenses"
                                            ] = total_research_and_development_expenses_list[
                                                i
                                            ]
                                            json_data[
                                                "total_other_expenses"
                                            ] = total_other_expenses_list[i]

                                            # print(json_data)

                                            r = requests.post(
                                                base_url + "v1/opex/",
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)

                                    # If opex expenses data is available for that year
                                    else:

                                        opex_expenses_table_result = json.loads(
                                            opex_expenses_table_result.text
                                        )
                                        opex_expenses_table_result = pd.DataFrame(
                                            opex_expenses_table_result
                                        )

                                        try:
                                            opex_expenses_table_result = (
                                                opex_expenses_table_result.sort_values(
                                                    by="month", ascending=True
                                                )
                                            )
                                        except:
                                            # print("Block1")
                                            return jsonify([])

                                        opex_expenses_table_result.index = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]
                                        opex_expenses_table_result = (
                                            opex_expenses_table_result.where(
                                                opex_expenses_table_result.notnull(),
                                                None,
                                            )
                                        )

                                        # print(opex_expenses_table_result)

                                        opex_expenses_table_result[
                                            "total_general_and_administrative_expenses"
                                        ] = opex_expenses_data_table[
                                            "Total General and Administrative Expenses"
                                        ]
                                        opex_expenses_table_result[
                                            "total_sales_and_marketing_expenses"
                                        ] = opex_expenses_data_table[
                                            "Total Sales and Marketing Expenses"
                                        ]
                                        opex_expenses_table_result[
                                            "total_research_and_development_expenses"
                                        ] = opex_expenses_data_table[
                                            "Total Research and Development Expenses"
                                        ]
                                        opex_expenses_table_result[
                                            "total_other_expenses"
                                        ] = opex_expenses_data_table[
                                            "Total Other Expenses"
                                        ]

                                        opex_expenses_data = opex_expenses_table_result

                                        # print(opex_expenses_data)

                                        opex_expenses_data = opex_expenses_data.to_dict(
                                            orient="records"
                                        )

                                        opex_expenses_data = json.dumps(
                                            opex_expenses_data
                                        )

                                        opex_range = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        id_list = opex_expenses_table_result[
                                            "_id"
                                        ].tolist()

                                        # print(id_list)

                                        total_general_and_administrative_expenses_list = opex_expenses_table_result[
                                            "total_general_and_administrative_expenses"
                                        ].tolist()
                                        total_sales_and_marketing_expenses_list = (
                                            opex_expenses_table_result[
                                                "total_sales_and_marketing_expenses"
                                            ].tolist()
                                        )
                                        total_research_and_development_expenses_list = opex_expenses_table_result[
                                            "total_research_and_development_expenses"
                                        ].tolist()
                                        total_other_expenses_list = (
                                            opex_expenses_table_result[
                                                "total_other_expenses"
                                            ].tolist()
                                        )

                                        json_data = {
                                            "total_general_and_administrative_expenses": 0,
                                            "total_sales_and_marketing_expenses": 0,
                                            "total_research_and_development_expenses": 0,
                                            "total_other_expenses": 0,
                                        }

                                        for i in opex_range:

                                            json_data[
                                                "total_general_and_administrative_expenses"
                                            ] = total_general_and_administrative_expenses_list[
                                                i
                                            ]
                                            json_data[
                                                "total_sales_and_marketing_expenses"
                                            ] = total_sales_and_marketing_expenses_list[
                                                i
                                            ]

                                            json_data[
                                                "total_research_and_development_expenses"
                                            ] = total_research_and_development_expenses_list[
                                                i
                                            ]
                                            json_data[
                                                "total_other_expenses"
                                            ] = total_other_expenses_list[i]

                                            id = id_list[i]

                                            # print(json_data)

                                            r = requests.put(
                                                base_url + "v1/opex/{}".format(id),
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )

                                            # print(json_data)

                                    print("Opex Expenses Sheet Processed")

                                elif sheet_list[index] == "Employee":
                                    # check for Employee data
                                    employee_data = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Employee",
                                        usecols=["Table Name", "Year"],
                                        dtype={
                                            "Table Name": str,
                                        },
                                    )

                                    # print(employee_data)
                                    # print(employee_data["Table Name"][0])

                                    employee_data_table = pd.read_excel(
                                        request.files["file"],
                                        sheet_name="Employee",
                                        usecols=[
                                            "Employee Name",
                                            "Department",
                                            "Role Type",
                                            "Role Name",
                                            "CXO",
                                            "Annual Salary",
                                            "Start Date",
                                            "End Date",
                                        ],
                                        dtype={
                                            "Employee Name": str,
                                            "Department": str,
                                            "Role Type": str,
                                            "Role Name": str,
                                            "CXO": bool,
                                            "Annual Salary": float,
                                            "Start Date": date,
                                            "End Date": date,
                                        },
                                    )

                                    employee_data_table = employee_data_table.where(
                                        employee_data_table.notnull(), None
                                    )

                                    # print(employee_data_table)

                                    employee_name_list = employee_data_table[
                                        "Employee Name"
                                    ].tolist()

                                    for i in employee_name_list:
                                        employee_name = str(i)
                                        # print(employee_name)

                                        # Check if employee exists

                                        employee_table_result = requests.get(
                                            base_url
                                            + "v1/employee?"
                                            + "page={}&page_size={}&startup_id={}&employee_name={}".format(
                                                page,
                                                page_size,
                                                startup_id,
                                                employee_name,
                                            ),
                                            headers={
                                                "Content-Type": "application/json",
                                                "Authorization": "Bearer {}".format(
                                                    access_token
                                                ),
                                            },
                                        )

                                        # print(employee_table_result.text)

                                        # If employee does not exist, create employee

                                        if len(employee_table_result.text) < 4:

                                            # print(
                                            #     "Employee "
                                            #     + employee_name
                                            #     + " does not exist"
                                            # )

                                            json_data = {
                                                "annual_salary": None,
                                                "cxo": None,
                                                "department": "",
                                                "employee_id": "",
                                                "employee_name": "",
                                                "role_name": "",
                                                "role_type": "",
                                                "start_date": None,
                                                "end_data": None,
                                                "startup_id": "",
                                            }

                                            json_data["startup_id"] = str(startup_id)
                                            json_data["employee_name"] = str(
                                                employee_name
                                            )
                                            json_data["annual_salary"] = float(
                                                employee_data_table["Annual Salary"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["cxo"] = bool(
                                                employee_data_table["CXO"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["department"] = str(
                                                employee_data_table["Department"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["role_name"] = str(
                                                employee_data_table["Role Name"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["role_type"] = str(
                                                employee_data_table["Role Type"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["start_date"] = str(
                                                (
                                                    employee_data_table["Start Date"][
                                                        int(
                                                            employee_data_table[
                                                                employee_data_table[
                                                                    "Employee Name"
                                                                ]
                                                                == employee_name
                                                            ].index.values
                                                        )
                                                    ]
                                                    .to_pydatetime()
                                                    .date()
                                                )
                                            )

                                            date_check = employee_data_table[
                                                "End Date"
                                            ][
                                                int(
                                                    employee_data_table[
                                                        employee_data_table[
                                                            "Employee Name"
                                                        ]
                                                        == employee_name
                                                    ].index.values
                                                )
                                            ]

                                            if date_check == None:

                                                json_data["end_data"] = None

                                            else:
                                                json_data["end_data"] = str(
                                                    (
                                                        employee_data_table["End Date"][
                                                            int(
                                                                employee_data_table[
                                                                    employee_data_table[
                                                                        "Employee Name"
                                                                    ]
                                                                    == employee_name
                                                                ].index.values
                                                            )
                                                        ]
                                                        .to_pydatetime()
                                                        .date()
                                                    )
                                                )

                                            # print(json_data)

                                            r = requests.post(
                                                base_url + "v1/employee/",
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )
                                            # print(json_data)

                                        # If employee exists, update employee

                                        else:
                                            # print(
                                            #     "Employee " + employee_name + " exist"
                                            # )

                                            employee_table_result = json.loads(
                                                employee_table_result.text
                                            )

                                            employee_table_result = pd.DataFrame(
                                                employee_table_result
                                            )

                                            id = employee_table_result["_id"].to_list()[
                                                0
                                            ]

                                            # print(id)

                                            json_data = {
                                                "annual_salary": None,
                                                "cxo": None,
                                                "department": "",
                                                "role_name": "",
                                                "role_type": "",
                                                "start_date": None,
                                                "end_data": None,
                                            }

                                            json_data["annual_salary"] = float(
                                                employee_data_table["Annual Salary"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["cxo"] = bool(
                                                employee_data_table["CXO"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["department"] = str(
                                                employee_data_table["Department"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["role_name"] = str(
                                                employee_data_table["Role Name"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["role_type"] = str(
                                                employee_data_table["Role Type"][
                                                    int(
                                                        employee_data_table[
                                                            employee_data_table[
                                                                "Employee Name"
                                                            ]
                                                            == employee_name
                                                        ].index.values
                                                    )
                                                ]
                                            )

                                            json_data["start_date"] = str(
                                                (
                                                    employee_data_table["Start Date"][
                                                        int(
                                                            employee_data_table[
                                                                employee_data_table[
                                                                    "Employee Name"
                                                                ]
                                                                == employee_name
                                                            ].index.values
                                                        )
                                                    ]
                                                    .to_pydatetime()
                                                    .date()
                                                )
                                            )

                                            date_check = employee_data_table[
                                                "End Date"
                                            ][
                                                int(
                                                    employee_data_table[
                                                        employee_data_table[
                                                            "Employee Name"
                                                        ]
                                                        == employee_name
                                                    ].index.values
                                                )
                                            ]

                                            if date_check == None:

                                                json_data["end_data"] = None

                                            else:
                                                json_data["end_data"] = str(
                                                    (
                                                        employee_data_table["End Date"][
                                                            int(
                                                                employee_data_table[
                                                                    employee_data_table[
                                                                        "Employee Name"
                                                                    ]
                                                                    == employee_name
                                                                ].index.values
                                                            )
                                                        ]
                                                        .to_pydatetime()
                                                        .date()
                                                    )
                                                )

                                            r = requests.put(
                                                base_url + "v1/employee/{}".format(id),
                                                data=json.dumps(json_data),
                                                headers={
                                                    "Content-Type": "application/json",
                                                    "Authorization": "Bearer {}".format(
                                                        access_token
                                                    ),
                                                },
                                            )

                                            # print(json_data)

                                    print("Employee Sheet Processed")

                                elif sheet_list[index] == "Product":
                                    # check for the startup
                                    if (
                                        startup_id
                                        == "Tyke Technologies Private Limited"
                                        or startup_id == "startup-1slug"
                                    ):
                                        product_data = pd.read_excel(
                                            request.files["file"],
                                            sheet_name="Product",
                                            usecols=["Table Name", "Year"],
                                            dtype={
                                                "Table Name": str,
                                            },
                                        )
                                        product_data["Year"] = product_data[
                                            "Year"
                                        ].fillna(0)
                                        product_data = product_data.astype(
                                            {"Year": "int"}
                                        )
                                        year = product_data["Year"][0]

                                        # print(product_data)
                                        # print(product_data["Table Name"][0])
                                        # print(product_data["Year"][0])

                                        product_data_table = pd.read_excel(
                                            request.files["file"],
                                            sheet_name="Product",
                                            usecols=[
                                                "Table Column Name",
                                                "January",
                                                "February",
                                                "March",
                                                "April",
                                                "May",
                                                "June",
                                                "July",
                                                "August",
                                                "September",
                                                "October",
                                                "November",
                                                "December",
                                            ],
                                            dtype={
                                                "Table Column Name": str,
                                            },
                                        )

                                        product_data_table = product_data_table.where(
                                            product_data_table.notnull(), None
                                        )

                                        # print(product_data_table)

                                        product_data_table = (
                                            product_data_table.set_index(
                                                "Table Column Name"
                                            )
                                        )
                                        product_data_table.index.names = [None]
                                        product_data_table = (
                                            product_data_table.transpose()
                                        )

                                        product_data_table = pd.DataFrame(
                                            product_data_table, dtype=object
                                        )

                                        product_data_table.index = [
                                            0,
                                            1,
                                            2,
                                            3,
                                            4,
                                            5,
                                            6,
                                            7,
                                            8,
                                            9,
                                            10,
                                            11,
                                        ]

                                        # print(product_data_table)

                                    # Check if product data is available for that year

                                    product_table_result = requests.get(
                                        base_url
                                        + "v1/product?"
                                        + "page={}&page_size={}&startup_id={}&year={}".format(
                                            page, page_size, startup_id, year
                                        ),
                                        headers={
                                            "Content-Type": "application/json",
                                            "Authorization": "Bearer {}".format(
                                                access_token
                                            ),
                                        },
                                    )

                                    # If product data is not available for that year
                                    if len(product_table_result.text) < 4:

                                        json_data = {
                                            "startup_id": "",
                                            "year": 0,
                                            "dataset": [[], [], [], [], []],
                                            "labels": [[], [], [], [], []],
                                        }

                                        json_data["startup_id"] = str(startup_id)
                                        json_data["year"] = int(year)
                                        json_data["dataset"][0] = product_data_table[
                                            "No. of Campaigns"
                                        ].to_list()
                                        json_data["dataset"][1] = product_data_table[
                                            "Investors Participated"
                                        ].to_list()
                                        json_data["dataset"][2] = product_data_table[
                                            "Total Invested Amount"
                                        ].to_list()
                                        json_data["dataset"][3] = [
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                        ]
                                        json_data["dataset"][4] = [
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                        ]

                                        json_data["labels"] = [
                                            "No. of Campaigns",
                                            "Investors Participated",
                                            "Total Invested Amount",
                                            "Avg Investor Participation per Campaign",
                                            "Avg Investor Investment Amount per Campaign",
                                        ]

                                        # print(json_data)

                                        r = requests.post(
                                            base_url + "v1/product/",
                                            data=json.dumps(json_data),
                                            headers={
                                                "Content-Type": "application/json",
                                                "Authorization": "Bearer {}".format(
                                                    access_token
                                                ),
                                            },
                                        )

                                        # print(json_data)

                                    # If product data is available for that year
                                    else:

                                        product_table_result = json.loads(
                                            product_table_result.text
                                        )

                                        # print(product_table_result)
                                        # print(product_table_result[0]["_id"])
                                        # print(product_table_result[0]["dataset"][0])
                                        # print(product_table_result[0]["dataset"][1])
                                        # print(product_table_result[0]["dataset"][2])
                                        # print(product_table_result[0]["labels"][0])
                                        # print(product_table_result[0]["labels"][1])
                                        # print(product_table_result[0]["labels"][2])

                                        product_table_result[0]["dataset"][
                                            0
                                        ] = product_data_table[
                                            "No. of Campaigns"
                                        ].to_list()

                                        product_table_result[0]["dataset"][
                                            1
                                        ] = product_data_table[
                                            "Investors Participated"
                                        ].to_list()

                                        product_table_result[0]["dataset"][
                                            2
                                        ] = product_data_table[
                                            "Total Invested Amount"
                                        ].to_list()

                                        # print(product_table_result[0]["_id"])
                                        # print(product_table_result[0]["dataset"][0])
                                        # print(product_table_result[0]["dataset"][1])
                                        # print(product_table_result[0]["dataset"][2])
                                        # print(product_table_result[0]["labels"][0])
                                        # print(product_table_result[0]["labels"][1])
                                        # print(product_table_result[0]["labels"][2])

                                        id = product_table_result[0]["_id"]

                                        json_data = product_table_result[0]

                                        r = requests.put(
                                            base_url + "v1/product/{}".format(id),
                                            data=json.dumps(json_data),
                                            headers={
                                                "Content-Type": "application/json",
                                                "Authorization": "Bearer {}".format(
                                                    access_token
                                                ),
                                            },
                                        )

                                        # print(json_data)

                                    print("Product Sheet Processed")

                                else:
                                    print("Startup does not have a product table")

                                index += 1

                            return "File uploaded"
